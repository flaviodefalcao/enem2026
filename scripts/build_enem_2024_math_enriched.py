#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import fitz
import pandas as pd
from openpyxl import load_workbook


COMPETENCE_THEME_MAP = {
    1: {
        "theme": "Conhecimentos numéricos",
        "subtheme": "Números, operações e contagem",
    },
    2: {
        "theme": "Conhecimentos geométricos",
        "subtheme": "Geometria plana e espacial",
    },
    3: {
        "theme": "Grandezas e medidas",
        "subtheme": "Escalas, medidas e proporcionalidade geométrica",
    },
    4: {
        "theme": "Variação de grandezas",
        "subtheme": "Relações de dependência e proporcionalidade",
    },
    5: {
        "theme": "Conhecimentos algébricos",
        "subtheme": "Funções, equações e modelagem algébrica",
    },
    6: {
        "theme": "Gráficos e tabelas",
        "subtheme": "Leitura, interpretação e inferência com dados",
    },
    7: {
        "theme": "Estatística e probabilidade",
        "subtheme": "Medidas estatísticas, amostragem e chance",
    },
}

COMPETENCE_SUBTHEME_RULES = {
    1: [
        {
            "subtheme": "Combinatória e contagem",
            "statement_patterns": [
                r"\bnúmero máximo de maneiras\b",
                r"\bequipe com\b",
                r"\bsortead[oa]\b",
                r"\bprincípios? de contagem\b",
            ],
            "fallback_patterns": [r"\bcontagem\b"],
        },
        {
            "subtheme": "Números racionais e equivalência",
            "statement_patterns": [
                r"\bnúmeros racionais\b",
                r"\bmesma quantidade\b",
                r"\brepresentações? de números racionais\b",
                r"\bfraç",
                r"\bdecimal\b",
            ],
        },
    ],
    2: [
        {
            "subtheme": "Geometria plana e espacial",
            "statement_patterns": [
                r"\btriângulo\b",
                r"\bquadrado\b",
                r"\bhexágono\b",
                r"\bcilindro\b",
                r"\bvolume\b",
                r"\bárea\b",
                r"\bprojeção ortogonal\b",
                r"\bcircunfer",
                r"\btrapézio\b",
                r"\bequidistantes?\b",
                r"\bmesmo plano\b",
                r"\besquema\b",
            ],
            "fallback_patterns": [r"\bespaço e forma\b"],
        }
    ],
    3: [
        {
            "subtheme": "Medidas, escalas e unidades",
            "statement_patterns": [
                r"\bmicrômetros?\b",
                r"\bmilímetros?\b",
                r"\bpolegadas?\b",
                r"\bmetros?\b",
                r"\bquilômetros?\b",
                r"\blitros?\b",
                r"\bcapacidade\b",
                r"\bmedida\b",
                r"\bescala\b",
                r"\bdensidade demográfica\b",
            ],
        }
    ],
    4: [
        {
            "subtheme": "Relações de proporcionalidade",
            "statement_patterns": [
                r"\bporcent",
                r"\btrês quartos\b",
                r"\bmetade\b",
                r"\brazão\b",
                r"\bproporção\b",
                r"\bproporções\b",
                r"\bdiretamente proporcional\b",
                r"\binversamente proporcionais?\b",
                r"\b40%\b",
                r"\b30% a menos\b",
            ],
            "fallback_patterns": [r"\bvariação de grandezas\b"],
        }
    ],
    5: [
        {
            "subtheme": "Funções e interpretação de gráficos",
            "statement_patterns": [
                r"\bgráfico\b",
                r"\bgráficos\b",
                r"\breta de tendência\b",
                r"\bsenoide\b",
                r"\boscila",
                r"\bgráfico cartesiano\b",
            ],
            "fallback_patterns": [r"\bfunção\b", r"\bfunções\b"],
        },
        {
            "subtheme": "Modelagem algébrica",
            "statement_patterns": [
                r"\bexpressão algébrica\b",
                r"\bexpressão numérica\b",
                r"\bequação\b",
                r"\bmodelagem\b",
                r"\balgébr",
                r"\bcifra de césar\b",
                r"\blog\b",
                r"\bvaria em função de\b",
                r"\ba expressão que fornece\b",
            ],
            "fallback_patterns": [r"\brepresentações? algébricas\b"],
        },
    ],
    6: [
        {
            "subtheme": "Leitura, interpretação e inferência com dados",
            "statement_patterns": [
                r"\bgráfico\b",
                r"\btabela\b",
                r"\bgráficos\b",
                r"\btabelas\b",
                r"\binfer",
            ],
        }
    ],
    7: [
        {
            "subtheme": "Medidas estatísticas, amostragem e chance",
            "statement_patterns": [
                r"\bprobabilidade\b",
                r"\bdistribuição estatística\b",
                r"\bmediana\b",
                r"\bmédia mensal\b",
                r"\bamostra",
            ],
        }
    ],
}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Enriquece a base de Matemática do ENEM 2024 com itens, matriz e dicionário."
    )
    parser.add_argument("--content-json", required=True)
    parser.add_argument("--items-csv", required=True)
    parser.add_argument("--matrix-pdf", required=True)
    parser.add_argument("--dictionary-xlsx", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--color", default="AZUL")
    parser.add_argument("--proof", type=int, default=None)
    return parser


def normalize_spaces(value: str) -> str:
    return " ".join(value.replace("\n", " ").split()).strip()


def classify_theme(
    statement: str,
    skill_description: str,
    competence_number: int,
) -> dict[str, str]:
    theme_info = COMPETENCE_THEME_MAP[competence_number]
    statement_haystack = normalize_spaces(statement).lower()
    fallback_haystack = normalize_spaces(f"{statement} {skill_description}").lower()

    for rule in COMPETENCE_SUBTHEME_RULES.get(competence_number, []):
        if any(
            re.search(pattern, statement_haystack, flags=re.IGNORECASE)
            for pattern in rule.get("statement_patterns", [])
        ):
            return {"theme": theme_info["theme"], "subtheme": rule["subtheme"]}

    for rule in COMPETENCE_SUBTHEME_RULES.get(competence_number, []):
        if any(
            re.search(pattern, fallback_haystack, flags=re.IGNORECASE)
            for pattern in rule.get("fallback_patterns", [])
        ):
            return {"theme": theme_info["theme"], "subtheme": rule["subtheme"]}

    return theme_info


def load_content_payload(content_json_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    payload = json.loads(content_json_path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return payload, {}
    if isinstance(payload, dict) and isinstance(payload.get("questions"), list):
        return payload["questions"], payload.get("metadata", {})
    raise SystemExit("Formato inválido de content-json. Esperado lista ou objeto com `questions`.")


def parse_math_matrix(matrix_pdf_path: Path) -> dict[str, Any]:
    pdf = fitz.open(matrix_pdf_path)
    math_pages_text = "\n".join(pdf[i].get_text("text") for i in [4, 5, 6])
    object_page_blocks = pdf[15].get_text("blocks")

    section_pattern = re.compile(
        r"Competência de área\s+(\d+)\s*-\s*(.+?)(?=\n\s*Competência de área\s+\d+\s*-|\Z)",
        re.S,
    )
    skill_pattern = re.compile(
        r"H(\d+)\s*-\s*(.+?)(?=\nH\d+\s*-|\Z)",
        re.S,
    )

    competencies: dict[int, dict[str, Any]] = {}
    for section_match in section_pattern.finditer(math_pages_text):
        competence_number = int(section_match.group(1))
        section_text = section_match.group(2)
        lines = section_text.splitlines()
        competence_description_lines = []
        skill_start_index = 0

        for index, line in enumerate(lines):
            if re.match(r"\s*H\d+\s*-", line):
                skill_start_index = index
                break
            competence_description_lines.append(line)

        competencies[competence_number] = {
            "competenceNumber": competence_number,
            "competenceDescription": normalize_spaces(
                "\n".join(competence_description_lines)
            ),
            "skills": {},
        }

        skills_text = "\n".join(lines[skill_start_index:])
        for skill_match in skill_pattern.finditer(skills_text):
            skill_number = int(skill_match.group(1))
            competencies[competence_number]["skills"][skill_number] = {
                "skillNumber": skill_number,
                "skillCode": f"H{skill_number}",
                "skillDescription": normalize_spaces(skill_match.group(2)),
            }

    objects_of_knowledge = []
    for block in object_page_blocks:
      text = normalize_spaces(block[4])
      if text.startswith("• "):
          objects_of_knowledge.append(text[2:])

    return {
        "competencies": competencies,
        "objectsOfKnowledge": objects_of_knowledge,
    }


def parse_dictionary(dictionary_xlsx_path: Path) -> dict[str, Any]:
    workbook = load_workbook(dictionary_xlsx_path, read_only=True, data_only=True)
    sheet = workbook["ITENS_PROVA_2024"]

    used_fields = {
        "CO_POSICAO",
        "SG_AREA",
        "CO_ITEM",
        "TX_GABARITO",
        "CO_HABILIDADE",
        "TX_COR",
        "CO_PROVA",
    }

    variables: dict[str, dict[str, Any]] = {}
    current_variable = None

    for row in sheet.iter_rows(values_only=True):
        values = list(row[:6])
        if not any(value is not None for value in values):
            continue

        variable_name = values[0]
        description = values[1]
        category = values[2]
        category_description = values[3]
        size = values[4]
        data_type = values[5]

        if isinstance(variable_name, str) and variable_name in used_fields:
            current_variable = variable_name
            variables[current_variable] = {
                "name": current_variable,
                "description": normalize_spaces(str(description or "")),
                "size": size,
                "type": data_type,
                "categories": {},
            }
            if category is not None:
                variables[current_variable]["categories"][str(category)] = normalize_spaces(
                    str(category_description or "")
                )
            continue

        if isinstance(variable_name, str) and variable_name not in used_fields:
            current_variable = None
            continue

        if current_variable and variable_name is None and category is not None:
            variables[current_variable]["categories"][str(category)] = normalize_spaces(
                str(category_description or "")
            )

    return variables


def choose_proof_family(mt_items: pd.DataFrame, color: str) -> tuple[int, list[int]]:
    color_items = mt_items[mt_items["TX_COR"] == color].copy()
    if color_items.empty:
        raise SystemExit(f"Nenhuma prova encontrada para a cor {color}.")

    fingerprints: dict[tuple[tuple[int, int, str], ...], list[int]] = {}
    for proof, proof_df in color_items.groupby("CO_PROVA"):
        signature = tuple(
            (
                int(row.CO_POSICAO),
                int(row.CO_ITEM),
                str(row.TX_GABARITO),
            )
            for row in proof_df.sort_values("CO_POSICAO").itertuples()
        )
        fingerprints.setdefault(signature, []).append(int(proof))

    canonical_family = min(fingerprints.values(), key=lambda proofs: min(proofs))
    canonical_proof = min(canonical_family)
    return canonical_proof, sorted(canonical_family)


def choose_proof_by_booklet(
    mt_items: pd.DataFrame,
    color: str,
    booklet_number: int,
) -> tuple[int, list[int]] | None:
    color_items = mt_items[mt_items["TX_COR"] == color].copy()
    proofs = sorted(int(proof) for proof in color_items["CO_PROVA"].dropna().unique())

    exact_suffix_matches = [proof for proof in proofs if proof % 100 == booklet_number]
    if len(exact_suffix_matches) == 1:
        return exact_suffix_matches[0], [exact_suffix_matches[0]]

    last_digit_matches = [proof for proof in proofs if proof % 10 == booklet_number]
    if len(last_digit_matches) == 1:
        return last_digit_matches[0], [last_digit_matches[0]]

    return None


def build_enriched_questions(
    content: list[dict[str, Any]],
    items_df: pd.DataFrame,
    matrix: dict[str, Any],
    dictionary: dict[str, Any],
    canonical_proof: int,
    equivalent_proofs: list[int],
) -> list[dict[str, Any]]:
    math_items = (
        items_df[items_df["CO_PROVA"] == canonical_proof]
        .copy()
        .sort_values("CO_POSICAO")
    )

    content_by_position = {int(item["examQuestionNumber"]): item for item in content}
    enriched: list[dict[str, Any]] = []
    skill_to_competence = {
        skill_number: competence_number
        for competence_number, competence in matrix["competencies"].items()
        for skill_number in competence["skills"].keys()
    }

    for row in math_items.itertuples():
        position = int(row.CO_POSICAO)
        content_item = content_by_position.get(position)
        if not content_item:
            raise SystemExit(f"Questão {position} não encontrada no content-json.")

        skill_number = int(row.CO_HABILIDADE)
        competence_number = skill_to_competence[skill_number]
        competence = matrix["competencies"][competence_number]
        skill = competence["skills"][skill_number]
        theme_info = classify_theme(
            statement=str(content_item["statement"]),
            skill_description=skill["skillDescription"],
            competence_number=competence_number,
        )

        enriched.append(
            {
                **content_item,
                "coPosicao": position,
                "coItem": int(row.CO_ITEM),
                "coProva": canonical_proof,
                "coProvasEquivalentes": equivalent_proofs,
                "color": str(row.TX_COR),
                "areaCode": str(row.SG_AREA),
                "areaDescription": dictionary["SG_AREA"]["categories"].get(
                    str(row.SG_AREA), "Matemática"
                ),
                "gabarito": str(row.TX_GABARITO),
                "coHabilidade": skill_number,
                "skillCode": skill["skillCode"],
                "skillDescription": skill["skillDescription"],
                "competenceNumber": competence_number,
                "competenceDescription": competence["competenceDescription"],
                "theme": theme_info["theme"],
                "subtheme": theme_info["subtheme"],
                "knowledgeObjects": [theme_info["theme"]],
            }
        )

    return enriched


def main() -> None:
    args = build_parser().parse_args()

    content_json_path = Path(args.content_json).expanduser().resolve()
    items_csv_path = Path(args.items_csv).expanduser().resolve()
    matrix_pdf_path = Path(args.matrix_pdf).expanduser().resolve()
    dictionary_xlsx_path = Path(args.dictionary_xlsx).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    content, content_metadata = load_content_payload(content_json_path)
    items_df = pd.read_csv(items_csv_path, sep=";", encoding="latin-1")
    mt_items = items_df[items_df["SG_AREA"] == "MT"].copy()

    if args.proof is not None:
        canonical_proof = int(args.proof)
        equivalent_proofs = [canonical_proof]
    else:
        booklet_number = content_metadata.get("bookletNumber")
        proof_selection = None
        if booklet_number is not None:
            proof_selection = choose_proof_by_booklet(
                mt_items=mt_items,
                color=args.color,
                booklet_number=int(booklet_number),
            )

        if proof_selection is not None:
            canonical_proof, equivalent_proofs = proof_selection
        else:
            canonical_proof, equivalent_proofs = choose_proof_family(mt_items, args.color)

    matrix = parse_math_matrix(matrix_pdf_path)
    dictionary = parse_dictionary(dictionary_xlsx_path)
    enriched_questions = build_enriched_questions(
        content=content,
        items_df=mt_items,
        matrix=matrix,
        dictionary=dictionary,
        canonical_proof=canonical_proof,
        equivalent_proofs=equivalent_proofs,
    )

    payload = {
        "metadata": {
            "year": 2024,
            "areaCode": "MT",
            "areaDescription": "Matemática",
            "color": args.color,
            "canonicalProof": canonical_proof,
            "equivalentProofs": equivalent_proofs,
            "bookletNumber": content_metadata.get("bookletNumber"),
            "sourceFiles": {
                "contentJson": str(content_json_path),
                "itemsCsv": str(items_csv_path),
                "matrixPdf": str(matrix_pdf_path),
                "dictionaryXlsx": str(dictionary_xlsx_path),
            },
            "dictionaryFieldsUsed": dictionary,
            "mathObjectsOfKnowledge": matrix["objectsOfKnowledge"],
        },
        "questions": enriched_questions,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(
        f"Base enriquecida gerada com {len(enriched_questions)} questões em {output_path}"
    )


if __name__ == "__main__":
    main()
